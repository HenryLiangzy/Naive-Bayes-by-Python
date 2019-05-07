from openpyxl import load_workbook
import random

FILENAME = "Course_Project.xlsx"


def read_data(filename, start_point, end_point, sheet_number=1):
    train_data = load_workbook(filename)
    all_sheet = train_data.get_sheet_names()
    sheet = train_data.get_sheet_by_name(all_sheet[sheet_number-1])
    data =[]
    
    for row in sheet[start_point:end_point]:
        row_data = []
        for cell in row:
            row_data.append(cell.value)
        
        data.append(row_data)

    return data


def split_data(data_set):
    index = []

    while len(index) < 100:
        number = random.randrange(0, len(data_set))
        if number in index:
            continue
        else:
            index.append(number)

    index.sort(reverse=True)
    print(index)
    test_set = []
    for index_number in index:
        test_set.append(data_set[index_number])
        del data_set[index_number]
    
    return data_set, test_set


def merge_data(index, atrA, class_value=''):
    letters = 'ABCDEFGHIJ'
    new_string = str(letters[index]) + str(atrA) + str(class_value)

    return new_string


def training(data_set):
    train_model = {
        'A-1-1':0, 'A0-1':0, 'A1-1':0, 'A-10':0, 'A00':0, 'A10':0, 'A-11':0, 'A01':0, 'A11':0,
        'B-1-1':0, 'B0-1':0, 'B1-1':0, 'B-10':0, 'B00':0, 'B10':0, 'B-11':0, 'B01':0, 'B11':0,
        'C-1-1':0, 'C0-1':0, 'C1-1':0, 'C-10':0, 'C00':0, 'C10':0, 'C-11':0, 'C01':0, 'C11':0,
        'D-1-1':0, 'D0-1':0, 'D1-1':0, 'D-10':0, 'D00':0, 'D10':0, 'D-11':0, 'D01':0, 'D11':0,
        'E-1-1':0, 'E0-1':0, 'E1-1':0, 'E-10':0, 'E00':0, 'E10':0, 'E-11':0, 'E01':0, 'E11':0,
        'F-1-1':0, 'F0-1':0, 'F1-1':0, 'F-10':0, 'F00':0, 'F10':0, 'F-11':0, 'F01':0, 'F11':0,
        'G-1-1':0, 'G0-1':0, 'G1-1':0, 'G-10':0, 'G00':0, 'G10':0, 'G-11':0, 'G01':0, 'G11':0,
        'H-1-1':0, 'H1-1':0, 'H-10':0, 'H10':0, 'H-11':0, 'H11':0,
        'I0-1':0, 'I1-1':0, 'I00':0, 'I10':0, 'I01':0, 'I11':0,
        'J-1':0, 'J0':0, 'J1':0,
    }
    total_number = len(data_set)
    finish_number = 0

    for row in data_set:
        for index in range(9):
            key_intersect = merge_data(index, row[index], row[9])
            key_single = merge_data(index, row[index])
            if train_model.get(key_intersect, -1) == -1 and train_model.get(key_single, -1) != -1:
                train_model[key_intersect] = 0
            elif train_model.get(key_single, -1) == -1 and train_model.get(key_intersect, -1) != -1:
                train_model[key_single] = 0
            elif train_model.get(key_intersect, -1) == -1 and train_model.get(key_single, -1) == -1:
                train_model[key_intersect] = 0
                train_model[key_single] = 0
            
            train_model[key_intersect] += 1
            train_model[key_single] += 1

        train_model[merge_data(9, row[9])] += 1

        finish_number += 1
        print("Finish", finish_number, "data, total", round(finish_number*100/total_number, 2), "%")

    key_list = train_model.keys()

    for key in key_list:
        train_model[key] = train_model[key] / finish_number

    save_model(train_model)

    return train_model


def save_model(train_model):
    file = open("bayes_train_model.model", "w")
    file.write(str(train_model))
    file.close()

    return None


def load_model(file_path):
    file = open(file_path, 'r')
    data = file.read()
    train_model = eval(data)
    print("Load model successful")
    file.close()

    return train_model


def perdict(model, perdict_list):
    cl_0 = []
    cl_1 = []
    cl_m = []

    cl_single = []

    for index in range(9):
        key_0 = merge_data(index, perdict_list[index], 0)
        key_1 = merge_data(index, perdict_list[index], 1)
        key_m = merge_data(index, perdict_list[index], -1)

        cl_single.append(model[merge_data(index, perdict_list[index])])

        cl_0.append(model[key_0])
        cl_1.append(model[key_1])
        cl_m.append(model[key_m])

    mult_single = 1
    for ps in cl_single:
        mult_single = ps * mult_single
    
    # perdict class = 0
    p0 = model[merge_data(9, 0)]
    mult_0 = 1
    for p in cl_0:
        mult_0 = p / p0 * mult_0
    mult_0 = mult_0 * p0

    p_0 = mult_0 / mult_single

    # perdict class = 1
    p1 = model[merge_data(9, 1)]
    mult_1 = 1
    for p in cl_1:
        mult_1 = p / p1 * mult_1
    mult_1 = mult_1 * p1

    p_1 = mult_1 / mult_single

    # perdict class = -1
    pm = model[merge_data(9, -1)]
    mult_m = 1
    for p in cl_m:
        mult_m = p / pm * mult_m
    mult_m = mult_m * pm

    p_m = mult_m / mult_single

    # compare the value
    result = [p_0, p_1, p_m]
    index = result.index(max(result))

    if index == 0:
        return 0
    elif index == 1:
        return 1
    elif index == 2:
        return -1
    else:
        return None


def test_accuracy(model, test_list):
    total = len(test_list)
    correct = 0
    for row in test_list:
        condiction = []
        for index in range(9):
            condiction.append(row[index])
        
        perdict_res = perdict(model, condiction)

        if perdict_res == row[9]:
            correct += 1
    
    print("The accuracy is", round(correct/total, 4), "Correct:", correct, "Error:", total-correct)


def write_perdict(model, filename):
    perdict_data= read_data(filename, 'A3', 'I102', 2)
    train_data = load_workbook(filename)
    all_sheet = train_data.get_sheet_names()
    sheet = train_data.get_sheet_by_name(all_sheet[1])
    index = 3
    for row in perdict_data:
        result = perdict(model, row)
        sheet['J'+str(index)] = result
        print("Finish", index-2, "data")
        index += 1
    
    train_data.save("Course_Project_result.xlsx")

    return None



if __name__ == '__main__':
    data_set, test_set = split_data(read_data(FILENAME, 'A3', 'J1255', 1))
    train_model = training(data_set)
    test_accuracy(train_model, test_set)
    # train_model = load_model("bayes_train_model.model")
    write_perdict(train_model, FILENAME)